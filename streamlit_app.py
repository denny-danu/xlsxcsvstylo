#!/usr/bin/env python3
"""
Streamlit UI for XLSX → CSV product data pipeline.
Allows easy override of all key parameters and mappings.
"""

import streamlit as st
import csv
import json
import io
import openpyxl
from pathlib import Path
from loguru import logger

st.set_page_config(
    page_title="Product Data Pipeline",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] { background: #0f172a; }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stTextInput label,
    [data-testid="stSidebar"] .stNumberInput label { color: #94a3b8 !important; font-size: 0.8rem; }
    .main-header { font-size: 1.6rem; font-weight: 700; color: #1e293b; margin-bottom: 0; }
    .sub-header { color: #64748b; font-size: 0.9rem; margin-top: 0.2rem; margin-bottom: 1.5rem; }
    .stTab [data-baseweb="tab"] { font-size: 0.85rem; font-weight: 600; }
    .stDataFrame { font-size: 0.8rem; }
    div[data-testid="metric-container"] { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 0.5rem; }
    .stAlert { font-size: 0.85rem; }
    textarea { font-family: monospace !important; font-size: 0.8rem !important; }
    .brand-input { background: #fffbeb; border: 1px solid #fcd34d; border-radius: 6px; padding: 0.3rem 0.6rem; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# DEFAULT CONFIG
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_FIXED_VALUES = {
    "Brand":                              "Tonique",           # ← NEW
    "Category":                           "WOMEN > Clothing > T-shirts",
    "Designer Code":                      "TNQO",
    "Supplier":                           "ST-SUP-0097",
    "Purchase Type":                      "ASN",
    "Cancelation Terms / Return Policy":  "Try and Buy",
    "Max Cart Quantity":                  "6",
    "Quantity":                           "1",
    "Fitting":                            "Regular Fit",
    "Color":                              "Black",
    "First in First Out":                 "No",
    "New Arrival":                        "Yes",
    "Flash Drop":                         "Yes",
    "Limited Edition":                    "Yes",
    "Stylo Unique":                       "Yes",
    "Supplier Re-Order Threshold Quantity": "1000",
}

DEFAULT_FABRIC_MAP = {
    "jersey ultra dry": "Jersey Ultra Dry",
    "polyester": "Polyester", "polymicro": "Polyester",
    "micro print": "Polyester", "polly micro": "Polyester",
    "pollycotton": "Polyester", "spandex": "Spandex",
    "spandek": "Spandex", "rayon": "Rayon",
    "satin": "Satin", "silk": "Silky",
    "katun": "Cotton", "cotton": "Cotton",
    "linen": "Linen", "denim": "Denim",
    "viscose": "Viscose", "nylon": "Nylon",
    "wool": "Wool", "micro cotton": "Micro Cotton",
    "satin velvet": "Satin Velvet", "rayon cotton": "Rayon Cotton",
    "almeera silk": "Almeera Silk", "silky cotton": "Silky Cotton",
}

DEFAULT_SKU_FABRIC_MAP = {
    "cotton": "01C", "denim": "01D", "linen": "01L",
    "polyester": "01P", "rayon": "01R", "silk": "01S",
    "viscose": "01V", "wool": "01W", "pollycotton": "01T",
    "micro cotton": "01M", "satin velvet": "01V", "rayon cotton": "01X",
    "almeera silk": "01A", "silky cotton": "01Y", "jersey": "01P",
    "spandex": "01X", "jersey ultra dry": "01P", "polycarbonate": "01Z",
    "stainless steel": "02S", "tr90 flexible": "02T",
}

DEFAULT_SIZE_MAP = {
    "xs": "Extra Small", "s": "Small", "m": "Medium",
    "l": "Large", "xl": "Extra Large", "xxl": "2X Large",
    "2xl": "2X Large", "heavy": "One Size", "light": "One Size",
    "all size": "One Size",
}

DEFAULT_COLOR_MAP = {
    "anti slip black": "Black", "baby pink": "Soft Pink",
    "black/grey": "Black", "damier black": "Black",
    "nylon black": "Black", "premium black": "Black",
    "boost black": "Black Out", "damier baby pink": "Light Pink",
    "marble purple": "Purple", "terracotta": "Brick Red",
    "brick": "Brick Red", "tosca": "Aqua",
    "orange-mix": "Orange", "black-clear": "Black",
    "blue-mix": "Blue", "leopard": "Black", "black gold": "Black",
}

DEFAULT_CATEGORY_MAP = {
    "dress": "WOMEN > Clothing > Dresses",
    "t-shirt": "{PREFIX} > Clothing > T-shirts",
    "tshirt": "{PREFIX} > Clothing > T-shirts",
    "modest": "WOMEN > Clothing > Modest wear",
    "activewear": "{PREFIX} > Clothing > Active Wear",
    "sports": "{PREFIX} > Clothing > Active Wear",
    "wearable": "WOMEN > Watch & Accessories > Other Accessories",
    "accessory": "WOMEN > Watch & Accessories > Other Accessories",
    "bag": "MEN > Bags > Others",
    "eyewear": "{PREFIX} > Watch & Accessories > Sunglasses",
    "kacamata": "{PREFIX} > Watch & Accessories > Sunglasses",
    "sunglass": "{PREFIX} > Watch & Accessories > Sunglasses",
    "other": "KIDS > Kids Clothing > Other",
}

DEFAULT_SHEETS = [
    "Womens sports", "Mens sports", "Sports Accessories",
    "Mens Clothing", "Mens Watch & Accessories",
    "Womens Footwear", "Womens Clothing",
]

CSV_COLUMNS = [
    "Title", "Description", "SKU Code", "Barcode Number", "Image URL",
    "Digital Trial Image URL", "Digital Model Prompt", "Category", "Brand",
    "Designer Code", "Supplier", "Occasions", "Purchase Type",
    "Cancelation Terms / Return Policy", "Quantity", "Max Cart Quantity", "List Price",
    "Sell Price", "Gender", "Fitting", "Color", "Fabric", "Size",
    "First in First Out", "New Arrival", "Flash Drop", "Limited Edition",
    "Stylo Unique", "Tags", "Supplier Re-Order Threshold Quantity",
]

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════
if "fixed_values" not in st.session_state:
    st.session_state.fixed_values = DEFAULT_FIXED_VALUES.copy()
if "fabric_map" not in st.session_state:
    st.session_state.fabric_map = DEFAULT_FABRIC_MAP.copy()
if "sku_fabric_map" not in st.session_state:
    st.session_state.sku_fabric_map = DEFAULT_SKU_FABRIC_MAP.copy()
if "size_map" not in st.session_state:
    st.session_state.size_map = DEFAULT_SIZE_MAP.copy()
if "color_map" not in st.session_state:
    st.session_state.color_map = DEFAULT_COLOR_MAP.copy()
if "category_map" not in st.session_state:
    st.session_state.category_map = DEFAULT_CATEGORY_MAP.copy()
if "sheet_list" not in st.session_state:
    st.session_state.sheet_list = DEFAULT_SHEETS.copy()
if "shortcodes" not in st.session_state:
    st.session_state.shortcodes = {}

# Migrate: ensure Brand key exists for sessions that pre-date this change
if "Brand" not in st.session_state.fixed_values:
    st.session_state.fixed_values["Brand"] = DEFAULT_FIXED_VALUES["Brand"]

# ══════════════════════════════════════════════════════════════════════════════
# CORE LOGIC (using session state config)
# ══════════════════════════════════════════════════════════════════════════════

def map_gender(raw: str) -> str:
    if not raw: return "Female"
    r = raw.strip().lower()
    if r in ("woman","women","female","girl","girls"): return "Female"
    if r in ("man","men","male","boy","boys"): return "Male"
    return raw.strip()

def map_category(raw: str, gender: str = "Female", sheet_name: str = "") -> str:
    if not raw: return ""
    key = raw.strip().lower()
    prefix = "MEN" if gender == "Male" else "WOMEN"
    cmap = st.session_state.category_map
    for trigger, result in cmap.items():
        if trigger.lower() in key:
            return result.replace("{PREFIX}", prefix)
    return raw.strip()

def normalize_fabric(raw: str) -> str:
    if not raw: return ""
    rl = raw.strip().lower()
    fmap = st.session_state.fabric_map
    for k, v in fmap.items():
        if k.lower() in rl:
            return v
    return raw.strip()

def normalize_color(raw: str) -> str:
    if not raw: return ""
    rl = raw.strip().lower()
    cmap = st.session_state.color_map
    for k, v in cmap.items():
        if k.lower() in rl:
            return v
    if "grey" in rl: return "Grey"
    if "silver" in rl: return "Silver"
    if "white" in rl: return "White"
    return raw.strip().title()

def clean(val) -> str:
    if val is None: return ""
    if isinstance(val, float):
        try:
            if val == int(val): return str(int(val))
        except: pass
    return str(val).strip()

def normalize_brand(name: str) -> str:
    if not name: return ""
    b = name.strip()
    if "corenation" in b.lower(): return "Corenation"
    return b

def get_brand_code(brand_name: str) -> str:
    b = normalize_brand(brand_name).lower()
    sc = st.session_state.shortcodes
    return sc.get(b, (brand_name[:4]).upper().ljust(4,'X'))

def get_cat_abbr(cat_path: str) -> str:
    parts = [p.strip() for p in cat_path.split(">")]
    if len(parts) < 3: return (cat_path[:4]).upper().ljust(4,'X')
    return f"{parts[0][0].upper()}{parts[2][:3].upper()}"

def generate_sku(row: dict, size_abbr: str = None) -> str:
    brand_code = get_brand_code(str(row.get("Brand") or ""))
    cat_abbr = get_cat_abbr(str(row.get("Category") or ""))
    gender = "F" if str(row.get("Gender","")).lower() in ["female","woman","women","girls","girl"] else "M"
    fitting = "REG"
    color_raw = str(row.get("Color","")).lower()
    color = (color_raw[:3]).upper().ljust(3,'X')
    for k, code in [("white","WHT"),("black","BLK"),("grey","GRY"),("blue","BLU")]:
        if k in color_raw: color = code; break
    fabric_raw = str(row.get("Fabric","")).lower()
    sfmap = st.session_state.sku_fabric_map
    fabric = sfmap.get(fabric_raw, "01X")
    size_raw = (size_abbr or str(row.get("Size",""))).upper()
    size = "OS" if "ONE SIZE" in size_raw else size_raw
    title = (str(row.get("Title",""))[:4]).upper().ljust(4,'X')
    return f"{brand_code}-{cat_abbr}-{gender}-{fitting}-{color}-{fabric}-{size}-{title}"

def make_unique_sku(sku: str, tracker: dict) -> str:
    if sku not in tracker:
        tracker[sku] = 0; return sku
    tracker[sku] += 1
    return f"{sku}-{tracker[sku]:03d}"

def load_sheet(ws, sheet_name: str = "", sku_tracker: dict = None):
    if sku_tracker is None: sku_tracker = {}
    fv = st.session_state.fixed_values

    # ── Brand resolution: fixed value overrides sheet detection ──────────────
    fixed_brand = fv.get("Brand", "").strip()

    brand = ""
    if not fixed_brand:
        # Fall back to reading brand from the sheet header block
        for r in range(1, 11):
            for c in range(1, 6):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val and str(cell_val).strip().lower() == "brand":
                    brand = normalize_brand(clean(ws.cell(row=r, column=c+2).value))
                    break
            if brand: break
    else:
        brand = normalize_brand(fixed_brand)

    header_row = -1
    col_map = {}
    for r in range(1, 15):
        row_vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, 25)]
        if "Product Title" in row_vals:
            header_row = r
            for i, val in enumerate(row_vals):
                if not val: continue
                v = val.lower().replace(" ","").replace("/","").replace("(","").replace(")","").replace("-","")
                if v == "no": col_map["no"] = i
                elif "producttitle" in v: col_map["title"] = i
                elif "productdescription" in v: col_map["description"] = i
                elif v == "sku": col_map["sku"] = i
                elif "materials" in v: col_map["materials"] = i
                elif v == "category": col_map["category"] = i
                elif "gender" in v: col_map["gender"] = i
                elif "color" in v: col_map["color"] = i
                elif v == "size": col_map["size"] = i
                elif "price" in v: col_map["price"] = i
                elif "image" in v: col_map["image_link"] = i
            break

    if header_row == -1:
        return [], f"⚠️ Header row not found in sheet '{ws.title}'"

    rows = []
    current_parent = {}
    smap = st.session_state.size_map

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sku_raw = clean(r[col_map["sku"]]) if "sku" in col_map else ""
        if not sku_raw: continue

        no_val = r[col_map["no"]] if "no" in col_map else None
        title = clean(r[col_map["title"]]) if "title" in col_map else ""
        is_new_product = no_val is not None and title

        if is_new_product:
            gender_raw = clean(r[col_map.get("gender", -1)]) if "gender" in col_map else ""
            gender = map_gender(gender_raw)
            if not gender_raw and sheet_name:
                if "mens" in sheet_name.lower(): gender = "Male"
                elif "womens" in sheet_name.lower(): gender = "Female"
            current_parent = {
                "title": title,
                "description": clean(r[col_map["description"]]) if "description" in col_map else "",
                "category": map_category(clean(r[col_map["category"]]) if "category" in col_map else "", gender=gender, sheet_name=sheet_name),
                "gender": gender,
                "materials": normalize_fabric(clean(r[col_map["materials"]]) if "materials" in col_map else ""),
                "color": normalize_color(clean(r[col_map["color"]]) if "color" in col_map else ""),
                "brand": brand,
            }

        price = clean(r[col_map["price"]]) if "price" in col_map else ""
        size_raw = clean(r[col_map["size"]]) if "size" in col_map else ""
        color = normalize_color(clean(r[col_map["color"]]) if "color" in col_map else "") or current_parent.get("color","")

        output_row = {
            "Title":                               current_parent.get("title",""),
            "Description":                         current_parent.get("description",""),
            "SKU Code":                            sku_raw,
            "Barcode Number":                      "",
            "Image URL":                           "",
            "Digital Trial Image URL":             "",
            "Digital Model Prompt":                "",
            "Category":                            fv.get("Category", ""),
            "Brand":                               brand.capitalize(),
            "Designer Code":                       fv.get("Designer Code",""),
            "Supplier":                            fv.get("Supplier",""),
            "Occasions":                           "",
            "Purchase Type":                       fv.get("Purchase Type",""),
            "Cancelation Terms / Return Policy":   fv.get("Cancelation Terms / Return Policy",""),
            "Quantity":                            fv.get("Quantity") or current_parent.get("Qty") or 1,
            "Max Cart Quantity":                   fv.get("Max Cart Quantity",""),
            "List Price":                          price,
            "Sell Price":                          price,
            "Gender":                              current_parent.get("gender",""),
            "Fitting":                             fv.get("Fitting",""),
            "Color":                               fv.get("Color",""),
            "Fabric":                              current_parent.get("materials",""),
            "Size":                                smap.get(size_raw.lower(), size_raw),
            "First in First Out":                  fv.get("First in First Out",""),
            "New Arrival":                         fv.get("New Arrival",""),
            "Flash Drop":                          fv.get("Flash Drop",""),
            "Limited Edition":                     fv.get("Limited Edition",""),
            "Stylo Unique":                        fv.get("Stylo Unique",""),
            "Tags":                                "",
            "Supplier Re-Order Threshold Quantity": fv.get("Supplier Re-Order Threshold Quantity",""),
        }
        base_sku = generate_sku(output_row, size_abbr=size_raw)
        output_row["SKU Code"] = make_unique_sku(base_sku, sku_tracker)
        rows.append(output_row)

    return rows, None

# ══════════════════════════════════════════════════════════════════════════════
# HELPER: editable JSON dict widget
# ══════════════════════════════════════════════════════════════════════════════
def dict_editor(label: str, state_key: str, height: int = 250):
    current = st.session_state[state_key]
    txt = st.text_area(
        label,
        value=json.dumps(current, indent=2),
        height=height,
        key=f"_ta_{state_key}",
        help="Edit as JSON. Changes apply on next run.",
    )
    try:
        parsed = json.loads(txt)
        if isinstance(parsed, dict):
            st.session_state[state_key] = parsed
            return True, None
        else:
            st.error("Must be a JSON object (dict).")
            return False, "Not a dict"
    except json.JSONDecodeError as e:
        st.error(f"JSON error: {e}")
        return False, str(e)

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Pipeline Config")
    st.caption("All settings persist during session.")

    st.markdown("---")
    st.markdown("**📋 Sheets to Process**")
    sheets_txt = st.text_area(
        "One sheet name per line",
        value="\n".join(st.session_state.sheet_list),
        height=160,
        key="_sheets_txt",
    )
    st.session_state.sheet_list = [s.strip() for s in sheets_txt.splitlines() if s.strip()]

    st.markdown("---")
    st.markdown("**🔑 Shortcodes JSON**")
    sc_txt = st.text_area(
        "Brand → Code mapping",
        value=json.dumps(st.session_state.shortcodes, indent=2),
        height=100,
        key="_sc_txt",
        placeholder='{"tonique": "TNQO", "corenation": "CORN"}',
    )
    try:
        sc_parsed = json.loads(sc_txt)
        if isinstance(sc_parsed, dict):
            st.session_state.shortcodes = {k.lower(): v for k, v in sc_parsed.items()}
    except: pass

    st.markdown("---")
    if st.button("↺ Reset All to Defaults", use_container_width=True):
        st.session_state.fixed_values = DEFAULT_FIXED_VALUES.copy()
        st.session_state.fabric_map = DEFAULT_FABRIC_MAP.copy()
        st.session_state.sku_fabric_map = DEFAULT_SKU_FABRIC_MAP.copy()
        st.session_state.size_map = DEFAULT_SIZE_MAP.copy()
        st.session_state.color_map = DEFAULT_COLOR_MAP.copy()
        st.session_state.category_map = DEFAULT_CATEGORY_MAP.copy()
        st.session_state.sheet_list = DEFAULT_SHEETS.copy()
        st.session_state.shortcodes = {}
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<p class="main-header">🏷️ Product Data Pipeline</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">XLSX → CSV converter with configurable mappings & fixed values</p>', unsafe_allow_html=True)

tab_run, tab_fixed, tab_maps, tab_sku = st.tabs(["▶ Run", "📌 Fixed Values", "🗺️ Mappings", "🔧 SKU Config"])

# ── TAB 1: RUN ─────────────────────────────────────────────────────────────────
with tab_run:
    col_up, col_info = st.columns([2, 1])
    with col_up:
        uploaded = st.file_uploader("Upload XLSX file", type=["xlsx"], label_visibility="collapsed")
    with col_info:
        if uploaded:
            st.success(f"✅ {uploaded.name}")
        else:
            st.info("Upload an XLSX file to begin")

    if uploaded:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), data_only=True)
        available_sheets = wb.sheetnames
        configured_sheets = st.session_state.sheet_list
        to_process = [s for s in configured_sheets if s in available_sheets]
        skipped = [s for s in configured_sheets if s not in available_sheets]

        c1, c2, c3 = st.columns(3)
        c1.metric("Available sheets", len(available_sheets))
        c2.metric("Configured sheets", len(configured_sheets))
        c3.metric("Will process", len(to_process))

        if skipped:
            st.warning(f"Sheets not found in file: {', '.join(skipped)}")

        with st.expander("Sheet list in uploaded file"):
            st.write(available_sheets)

        if st.button("🚀 Run Pipeline", type="primary", use_container_width=True):
            all_rows = []
            sku_tracker = {}
            log = []

            with st.spinner("Processing..."):
                for sheet_name in to_process:
                    rows, err = load_sheet(wb[sheet_name], sheet_name=sheet_name, sku_tracker=sku_tracker)
                    if err:
                        log.append(("warn", sheet_name, err))
                    else:
                        log.append(("ok", sheet_name, f"{len(rows)} rows"))
                        all_rows.extend(rows)

            for level, sheet, msg in log:
                if level == "ok":
                    st.success(f"**{sheet}**: {msg}")
                else:
                    st.warning(f"**{sheet}**: {msg}")

            if all_rows:
                import pandas as pd
                df = pd.DataFrame(all_rows, columns=CSV_COLUMNS)
                st.markdown(f"### Preview — {len(df)} rows")
                st.dataframe(df, use_container_width=True, height=350)

                buf = io.StringIO()
                writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS)
                writer.writeheader()
                writer.writerows(all_rows)

                output_name = Path(uploaded.name).stem + "_output.csv"
                st.download_button(
                    "⬇️ Download CSV",
                    data=buf.getvalue().encode("utf-8"),
                    file_name=output_name,
                    mime="text/csv",
                    use_container_width=True,
                    type="primary",
                )
            else:
                st.error("No rows were produced. Check sheet names and file structure.")

# ── TAB 2: FIXED VALUES ────────────────────────────────────────────────────────
with tab_fixed:
    st.markdown("These values are written as-is to every output row. Edit directly.")

    fv = st.session_state.fixed_values

    # ── Brand gets a prominent, highlighted input at the top ──────────────────
    st.markdown("**🏢 Brand** — applied to all rows; overrides any brand detected in the sheet")
    brand_val = st.text_input(
        "Brand name",
        value=fv.get("Brand", ""),
        key="fv_Brand",
        placeholder="e.g. Tonique",
    )
    st.session_state.fixed_values["Brand"] = brand_val
    if not brand_val.strip():
        st.warning("⚠️ Brand is empty — brand will be read from each sheet's header row instead.")

    st.markdown("---")

    # ── All other fixed values in a 2-column grid ─────────────────────────────
    other_keys = [k for k in fv.keys() if k != "Brand"]
    cols = st.columns(2)
    for i, key in enumerate(other_keys):
        with cols[i % 2]:
            new_val = st.text_input(key, value=fv[key], key=f"fv_{key}")
            st.session_state.fixed_values[key] = new_val

# ── TAB 3: MAPPINGS ────────────────────────────────────────────────────────────
with tab_maps:
    st.markdown("Edit lookup tables as JSON. Invalid JSON is ignored until fixed.")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Fabric Normalization**")
        dict_editor("raw input → display name", "fabric_map", height=220)
        st.markdown("**Color Normalization**")
        dict_editor("raw input → display name", "color_map", height=220)
    with c2:
        st.markdown("**Size Normalization**")
        dict_editor("abbr → full name", "size_map", height=150)
        st.markdown("**Category Mapping**")
        st.caption("Use `{PREFIX}` for gender-based prefix (WOMEN/MEN)")
        dict_editor("keyword in raw → category path", "category_map", height=280)

# ── TAB 4: SKU CONFIG ─────────────────────────────────────────────────────────
with tab_sku:
    st.markdown("**SKU Fabric Codes** — used in SKU string generation")
    st.caption("Format: `{ \"fabric name\": \"CODE\" }` e.g. `\"cotton\": \"01C\"`")
    dict_editor("fabric → SKU code", "sku_fabric_map", height=300)

    st.markdown("---")
    st.markdown("**SKU Format Preview**")
    st.code("{BRAND_CODE}-{CAT_ABBR}-{GENDER}-{FITTING}-{COLOR3}-{FABRIC_CODE}-{SIZE}-{TITLE4}")
    st.caption("BRAND_CODE from shortcodes.json · CAT_ABBR from category path · GENDER = F/M · FITTING = REG · COLOR3 = first 3 chars · FABRIC_CODE from table above · SIZE from size map · TITLE4 = first 4 chars of title")
